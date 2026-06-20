#!/usr/bin/env python3
"""
report.py
=========
Kolorowy raport z wynikow screenera (HTML; opcjonalnie XLSX).

Czyta CSV wyprodukowany przez main.py, grupuje spolki na
QUALITY / WATCHLIST / REJECT i pokazuje sygnaly per regula.

Uzycie:
    python report.py wyniki.csv                  # -> wyniki.html
    python report.py wyniki.csv -o raport.html
    python report.py wyniki.csv --xlsx raport.xlsx

Modul wystawia tez build_html() / write_xlsx(), by main.py mogl
wygenerowac raport jednym przebiegiem (flaga --report).
"""
from __future__ import annotations

import argparse
import csv
import html
from pathlib import Path

LABELS = ["quality", "watchlist", "reject", "error"]
LABEL_TITLE = {"quality": "QUALITY", "watchlist": "WATCHLIST",
               "reject": "REJECT", "error": "ERROR"}
LABEL_EMOJI = {"quality": "\U0001F7E2", "watchlist": "\U0001F7E1",
               "reject": "\U0001F534", "error": "\u26A0\uFE0F"}
LABEL_COLOR = {"quality": "#16a34a", "watchlist": "#d97706",
               "reject": "#dc2626", "error": "#6b7280"}

SIG_COLOR = {"green": "#16a34a", "warning": "#d97706", "red": "#dc2626", "na": "#94a3b8"}
SIG_EMOJI = {"green": "\U0001F7E2", "warning": "\u26A0\uFE0F",
             "red": "\U0001F534", "na": "\u2796"}

RULES = ["revenue_growth", "revenue_trend", "gross_margin", "gross_margin_trend",
         "operating_margin", "rule_of_40", "debt_to_revenue", "cash_runway",
         "valuation", "peg", "perf_6m", "perf_12m", "rsi", "pre_revenue", "insider"]
RULE_LABEL = {
    "revenue_growth": "Rev growth", "revenue_trend": "Rev trend",
    "gross_margin": "Gross margin", "gross_margin_trend": "GM trend",
    "operating_margin": "Op margin", "rule_of_40": "Rule of 40",
    "debt_to_revenue": "Debt/Rev", "cash_runway": "Runway",
    "valuation": "Wycena", "peg": "PEG", "perf_6m": "6M", "perf_12m": "12M",
    "rsi": "RSI", "pre_revenue": "Pre-rev", "insider": "Insider",
}

CSS = """
*{box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;
 background:#f1f5f9;color:#0f172a;margin:0;padding:28px;line-height:1.4}
h1{margin:0 0 4px;font-size:26px}
.sub{color:#475569;margin-bottom:6px;font-size:15px}
.legend{color:#64748b;font-size:13px;margin-bottom:18px}
.section{margin:26px 0 12px;font-size:19px;font-weight:700}
.card{background:#fff;border:1px solid #e2e8f0;border-radius:14px;padding:16px 18px;
 margin-bottom:12px;box-shadow:0 1px 3px rgba(15,23,42,.06)}
.cardhead{display:flex;align-items:center;gap:14px;flex-wrap:wrap}
.tk{font-size:18px;font-weight:700}
.pct{font-weight:700;font-size:16px}
.bar{height:8px;border-radius:6px;background:#e2e8f0;flex:1;min-width:140px;overflow:hidden}
.bar > span{display:block;height:100%}
.counts{font-size:13px;color:#475569;white-space:nowrap}
.metrics{display:flex;flex-wrap:wrap;gap:6px 18px;margin:12px 0;font-size:13px}
.metrics div{color:#64748b}
.metrics b{color:#0f172a}
.chips{display:flex;flex-wrap:wrap;gap:6px}
.chip{font-size:12px;padding:3px 9px;border-radius:999px;color:#fff;white-space:nowrap;cursor:default}
"""


def _num(row, key):
    v = row.get(key, "")
    if v in (None, "", "None", "nan", "NaN"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _pct(row, key):
    x = _num(row, key)
    return "\u2014" if x is None else f"{x * 100:.0f}%"


def _f2(row, key, suffix=""):
    x = _num(row, key)
    return "\u2014" if x is None else f"{x:.2f}{suffix}"


def _runway(row):
    x = _num(row, "cash_runway_months")
    if x is None:
        return "\u2014"
    return "\u221E" if x >= 999 else f"{x:.0f} mies."


def _esc(s):
    return html.escape(str(s if s is not None else ""))


def load_rows(csv_path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _metric_items(row):
    return [
        ("Rev YoY", _pct(row, "revenue_growth_yoy")),
        ("Gross margin", _pct(row, "gross_margin")),
        ("Op margin", _pct(row, "operating_margin")),
        ("Rule of 40", _f2(row, "rule_of_40")),
        ("Debt/Rev", _f2(row, "debt_to_revenue", "x")),
        ("Runway", _runway(row)),
        ("6M", _pct(row, "perf_6m")),
        ("12M", _pct(row, "perf_12m")),
        ("RSI-14", _f2(row, "rsi_14")),
        ("EV/Sales", _f2(row, "ev_to_sales", "x")),
        ("P/S", _f2(row, "ps_ratio", "x")),
        ("PEG", _f2(row, "peg_ratio")),
        ("Insider", _f2(row, "insider_net_ratio")),
    ]


def _chips_html(row):
    out = []
    for r in RULES:
        sig = (row.get(f"sig_{r}") or "na").strip().lower()
        if sig not in SIG_COLOR:
            sig = "na"
        reason = row.get(f"reason_{r}", "")
        out.append(
            f'<span class="chip" style="background:{SIG_COLOR[sig]}" '
            f'title="{_esc(reason)}">{SIG_EMOJI[sig]} {_esc(RULE_LABEL.get(r, r))}</span>'
        )
    return '<div class="chips">' + "".join(out) + "</div>"


def _card_html(row):
    label = (row.get("label") or "error").strip().lower()
    color = LABEL_COLOR.get(label, "#6b7280")
    pct = _num(row, "pct") or 0.0
    bar_w = max(0.0, min(100.0, pct))
    metrics = "".join(f"<div>{_esc(k)}: <b>{_esc(v)}</b></div>" for k, v in _metric_items(row))
    return (
        '<div class="card">'
        '<div class="cardhead">'
        f'<span class="tk">{LABEL_EMOJI.get(label, "")} {_esc(row.get("ticker", "?"))}</span>'
        f'<span class="pct" style="color:{color}">{pct:.1f}%</span>'
        f'<span class="bar"><span style="width:{bar_w:.0f}%;background:{color}"></span></span>'
        f'<span class="counts">\U0001F7E2 {_esc(row.get("greens", "0"))} &nbsp; '
        f'\u26A0\uFE0F {_esc(row.get("warnings", "0"))} &nbsp; '
        f'\U0001F534 {_esc(row.get("reds", "0"))} &nbsp; '
        f'\u2796 {_esc(row.get("na", "0"))}</span>'
        '</div>'
        f'<div class="metrics">{metrics}</div>'
        f'{_chips_html(row)}'
        '</div>'
    )


def build_html(rows, title="High-Growth Screener \u2014 raport"):
    counts = {}
    for r in rows:
        l = (r.get("label") or "error").strip().lower()
        counts[l] = counts.get(l, 0) + 1
    sub = " \u00B7 ".join(f"{LABEL_EMOJI[l]} {LABEL_TITLE[l]}: {counts[l]}"
                          for l in LABELS if counts.get(l))
    parts = [f"<h1>{_esc(title)}</h1>",
             f'<div class="sub">{len(rows)} spolek &nbsp;|&nbsp; {sub}</div>',
             '<div class="legend">Najedz kursorem na chip, aby zobaczyc uzasadnienie reguly. '
             'W grupach sortowanie wg wyniku malejaco.</div>']
    for l in LABELS:
        grp = [r for r in rows if (r.get("label") or "error").strip().lower() == l]
        if not grp:
            continue
        grp.sort(key=lambda r: _num(r, "pct") or 0.0, reverse=True)
        parts.append(f'<div class="section" style="color:{LABEL_COLOR[l]}">'
                     f'{LABEL_EMOJI[l]} {LABEL_TITLE[l]} ({len(grp)})</div>')
        parts.extend(_card_html(r) for r in grp)
    return ("<!doctype html><html lang='pl'><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            f"<title>{_esc(title)}</title><style>{CSS}</style></head>"
            f"<body>{''.join(parts)}</body></html>")


def write_html(rows, path, title="High-Growth Screener \u2014 raport"):
    Path(path).write_text(build_html(rows, title), encoding="utf-8")


def write_xlsx(rows, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise RuntimeError(
            "Eksport XLSX wymaga pakietu 'openpyxl'. Zainstaluj: pip install openpyxl "
            "(raport HTML dziala bez niego)."
        ) from e

    headers = ["ticker", "label", "pct", "greens", "warnings", "reds", "na",
               "revenue_growth_yoy", "gross_margin", "operating_margin", "rule_of_40",
               "debt_to_revenue", "cash_runway_months", "perf_6m", "perf_12m", "rsi_14",
               "ev_to_sales", "ps_ratio", "peg_ratio", "insider_net_ratio"]
    fills = {
        "quality": PatternFill("solid", fgColor="C6F6D5"),
        "watchlist": PatternFill("solid", fgColor="FEEBC8"),
        "reject": PatternFill("solid", fgColor="FED7D7"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Screener"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    srows = sorted(rows, key=lambda r: _num(r, "pct") or 0.0, reverse=True)
    for r in srows:
        ws.append([r.get(h, "") for h in headers])
        label = (r.get("label") or "").strip().lower()
        if label in fills:
            ws.cell(row=ws.max_row, column=2).fill = fills[label]
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 28)
    wb.save(path)


def main(argv=None):
    ap = argparse.ArgumentParser(description="Kolorowy raport z wynikow screenera")
    ap.add_argument("csv", help="Plik CSV z wynikami (z main.py)")
    ap.add_argument("-o", "--output", help="Sciezka raportu HTML (domyslnie obok CSV)")
    ap.add_argument("--xlsx", help="Dodatkowo zapisz raport XLSX pod ta sciezka")
    ap.add_argument("--title", default="High-Growth Screener \u2014 raport")
    a = ap.parse_args(argv)
    rows = load_rows(a.csv)
    out = a.output or str(Path(a.csv).with_suffix(".html"))
    write_html(rows, out, a.title)
    print(f"HTML  -> {out}  ({len(rows)} spolek)")
    if a.xlsx:
        try:
            write_xlsx(rows, a.xlsx)
            print(f"XLSX  -> {a.xlsx}")
        except RuntimeError as e:
            print(f"\u26A0\uFE0F  Pominieto XLSX: {e}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
